import streamlit as st
import pandas as pd
import os
from PIL import Image
import io
import base64
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage
import tempfile

# Sahifa konfiguratsiyasi
st.set_page_config(page_title="Omborxona Ma'lumotlari", layout="wide")

# Sessiyadagi o'zgaruvchilarni yaratish
if 'product_data' not in st.session_state:
    st.session_state.product_data = pd.DataFrame(columns=[
        'id', 'rasm', 'kod', 'toifa', 'davlat', 'dokon_id', 'omborchi', 'rang', 'olcham', 'miqdor', 'narx'
    ])

if 'temp_image' not in st.session_state:
    st.session_state.temp_image = None

if 'current_product_id' not in st.session_state:
    st.session_state.current_product_id = 1

# Toifa va boshqa konstantalar
TOIFALAR = ["Ayollar", "Erkaklar", "Bolalar", "Qizlar"]
RANGLAR = ["Qora", "Oq", "Qizil", "Ko'k", "Yashil", "Sariq", "Jigarrang", "Kulrang", "Pushti", "Binafsha"]
OLCHAMLAR = ["XS", "S", "M", "L", "XL", "XXL", "XXXL"]

# Rasimni kodga aylantirish
def image_to_base64(image):
    if image is None:
        return None
    
    buffered = io.BytesIO()
    image.save(buffered, format="PNG")
    return base64.b64encode(buffered.getvalue()).decode("utf-8")

# Mahsulot rasmini ko'rsatish
def display_image(image_data):
    if image_data:
        try:
            if isinstance(image_data, str) and image_data.startswith('data:image'):
                st.image(image_data)
            else:
                st.image(base64.b64decode(image_data))
        except Exception as e:
            st.error(f"Rasimni ko'rsatishda xatolik: {e}")
    else:
        st.info("Rasim yuklanmagan")

# Excel faylga saqlash funksiyasi
def save_to_excel(df):
    # Mahsulot ID va rasmli jadval
    products_unique = df[['id', 'rasm', 'kod', 'toifa']].drop_duplicates()
    
    # Toifa bo'yicha ma'lumotlar
    category_data = {}
    for category in TOIFALAR:
        category_data[category] = df[df['toifa'] == category]
    
    # Barcha ma'lumotlar
    wb = Workbook()
    
    # Birinchi sheet - hamma ma'lumotlar
    ws = wb.active
    ws.title = "Barcha_Mahsulotlar"
    
    # Ma'lumotlarni exceldagi sheetlarga yozish
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Har bir toifa uchun sheet
    for category in TOIFALAR:
        if not category_data[category].empty:
            ws = wb.create_sheet(title=category)
            for r in dataframe_to_rows(category_data[category], index=False, header=True):
                ws.append(r)
    
    # Mahsulot ID va rasmi uchun sheet
    ws = wb.create_sheet(title="Mahsulot_Rasmlari")
    ws.append(['id', 'kod', 'toifa', 'rasm'])
    
    row = 2
    for _, product in products_unique.iterrows():
        ws.cell(row=row, column=1, value=product['id'])
        ws.cell(row=row, column=2, value=product['kod'])
        ws.cell(row=row, column=3, value=product['toifa'])
        
        # Rasimni saqlash
        if product['rasm'] and not pd.isna(product['rasm']):
            try:
                with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as temp:
                    temp.write(base64.b64decode(product['rasm']))
                    temp_name = temp.name
                
                img = XLImage(temp_name)
                img.width = 100
                img.height = 100
                ws.add_image(img, f'D{row}')
                
                if os.path.exists(temp_name):
                    os.remove(temp_name)
            except Exception as e:
                ws.cell(row=row, column=4, value=f"Rasim xatosi: {str(e)}")
        row += 1
    
    # Faylni saqlash
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp:
        wb.save(temp.name)
        return temp.name

# Asosiy ilova
def main():
    st.title("Omborxona Ma'lumotlari")
    
    # Navbar - yuqori menu
    menu = ["Mahsulot qo'shish", "Mahsulotlar ro'yxati", "Excel yuklab olish"]
    choice = st.sidebar.selectbox("Menu", menu)
    
    # Umumiy ma'lumotlar
    with st.sidebar.expander("Umumiy ma'lumotlar"):
        common_dokon_id = st.text_input("Do'kon ID", "DOK001")
        common_omborchi = st.text_input("Omborchi ismi", "Abdullayev Abdulatif")
        common_davlat = st.text_input("Ishlab chiqarilgan davlat", "O'zbekiston")
    
    if choice == "Mahsulot qo'shish":
        st.subheader("Yangi mahsulot qo'shish")
        
        col1, col2 = st.columns(2)
        
        with col1:
            # Asosiy ma'lumotlar
            product_kod = st.text_input("Mahsulot kodi", "")
            product_toifa = st.selectbox("Toifa", TOIFALAR)
            
            # Rasmni yuklash
            uploaded_file = st.file_uploader("Mahsulot rasmini yuklang", type=['png', 'jpg', 'jpeg'])
            
            if uploaded_file is not None:
                try:
                    image = Image.open(uploaded_file)
                    st.session_state.temp_image = image_to_base64(image)
                    st.image(image, caption='Yuklangan rasim', use_column_width=True)
                except Exception as e:
                    st.error(f"Rasimni yuklashda xatolik: {e}")
        
        with col2:
            # Rangni tanlash va ranglar jadvalini qo'shish
            rangs = []
            with st.expander("Ranglar va o'lchamlarni qo'shish", expanded=True):
                num_colors = st.number_input("Nechta rang qo'shmoqchisiz?", min_value=1, max_value=10, value=1)
                
                for i in range(int(num_colors)):
                    st.subheader(f"Rang {i+1}")
                    color_row = {}
                    color_row['rang'] = st.selectbox(f"Rang {i+1}", RANGLAR, key=f"rang_{i}")
                    color_row['olcham'] = st.selectbox(f"O'lcham {i+1}", OLCHAMLAR, key=f"olcham_{i}")
                    color_row['miqdor'] = st.number_input(f"Miqdor {i+1}", min_value=0, value=1, key=f"miqdor_{i}")
                    color_row['narx'] = st.number_input(f"Narx {i+1} (so'm)", min_value=0, value=0, key=f"narx_{i}")
                    rangs.append(color_row)
            
        # Mahsulotni saqlash tugmasi
        if st.button("Mahsulotni saqlash"):
            if not product_kod:
                st.error("Mahsulot kodi kiritilishi shart!")
                return
            
            if st.session_state.temp_image is None:
                st.warning("Mahsulot rasmini yuklang!")
                return
            
            if not rangs:
                st.error("Kamida bitta rang qo'shish kerak!")
                return
            
            # Har bir rang uchun ma'lumotlarni qo'shish
            for rang_data in rangs:
                new_row = {
                    'id': st.session_state.current_product_id,
                    'rasm': st.session_state.temp_image,
                    'kod': product_kod,
                    'toifa': product_toifa,
                    'davlat': common_davlat,
                    'dokon_id': common_dokon_id,
                    'omborchi': common_omborchi,
                    'rang': rang_data['rang'],
                    'olcham': rang_data['olcham'],
                    'miqdor': rang_data['miqdor'],
                    'narx': rang_data['narx']
                }
                
                st.session_state.product_data = pd.concat([st.session_state.product_data, pd.DataFrame([new_row])], ignore_index=True)
            
            st.session_state.current_product_id += 1
            st.session_state.temp_image = None
            st.success("Mahsulot muvaffaqiyatli saqlandi!")
            st.experimental_rerun()
    
    elif choice == "Mahsulotlar ro'yxati":
        st.subheader("Mahsulotlar ro'yxati")
        
        if st.session_state.product_data.empty:
            st.info("Hozircha mahsulotlar yo'q. Mahsulotlarni 'Mahsulot qo'shish' sahifasidan qo'shing.")
        else:
            # Filtrlash
            filter_options = st.multiselect("Toifa bo'yicha filtrlash", TOIFALAR)
            
            filtered_data = st.session_state.product_data
            if filter_options:
                filtered_data = filtered_data[filtered_data['toifa'].isin(filter_options)]
            
            # Mahsulotlarni ID bo'yicha guruhlash
            unique_products = filtered_data[['id', 'kod', 'toifa']].drop_duplicates()
            
            for _, product in unique_products.iterrows():
                product_id = product['id']
                product_data = filtered_data[filtered_data['id'] == product_id]
                
                with st.expander(f"Mahsulot: {product['kod']} - {product['toifa']}"):
                    col1, col2 = st.columns([1, 2])
                    
                    with col1:
                        if not pd.isna(product_data['rasm'].iloc[0]):
                            display_image(product_data['rasm'].iloc[0])
                        else:
                            st.info("Rasim mavjud emas")
                    
                    with col2:
                        st.write(f"Kod: {product['kod']}")
                        st.write(f"Toifa: {product['toifa']}")
                        st.write(f"Davlat: {product_data['davlat'].iloc[0]}")
                        st.write(f"Do'kon ID: {product_data['dokon_id'].iloc[0]}")
                        st.write(f"Omborchi: {product_data['omborchi'].iloc[0]}")
                        
                        # Ranglar jadvali
                        st.subheader("Ranglar va o'lchamlar")
                        color_data = []
                        for _, row in product_data.iterrows():
                            color_data.append({
                                'Rang': row['rang'],
                                'O\'lcham': row['olcham'],
                                'Miqdor': row['miqdor'],
                                'Narx (so\'m)': row['narx']
                            })
                        
                        st.table(pd.DataFrame(color_data))
                    
                    # Tahrirlash va o'chirish tugmalari
                    col1, col2 = st.columns(2)
                    with col1:
                        if st.button(f"Tahrirlash #{product_id}", key=f"edit_{product_id}"):
                            st.info("Tahrirlash funksiyasi hozircha mavjud emas")
                    
                    with col2:
                        if st.button(f"O'chirish #{product_id}", key=f"delete_{product_id}"):
                            st.session_state.product_data = st.session_state.product_data[st.session_state.product_data['id'] != product_id]
                            st.success(f"Mahsulot (ID: {product_id}) muvaffaqiyatli o'chirildi")
                            st.experimental_rerun()
    
    elif choice == "Excel yuklab olish":
        st.subheader("Ma'lumotlarni Excel formatida yuklab olish")
        
        if st.session_state.product_data.empty:
            st.info("Hozircha mahsulotlar yo'q. Mahsulotlarni 'Mahsulot qo'shish' sahifasidan qo'shing.")
        else:
            if st.button("Excel faylni tayyorlash"):
                with st.spinner("Excel fayl tayyorlanmoqda..."):
                    excel_file = save_to_excel(st.session_state.product_data)
                    
                    with open(excel_file, "rb") as f:
                        bytes_data = f.read()

                    st.download_button(
                        label="Excel faylni yuklab olish",
                        data=bytes_data,
                        file_name="omborxona_malumotlari.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                    if os.path.exists(excel_file):
                        os.remove(excel_file)

if name == "main":
    main()
